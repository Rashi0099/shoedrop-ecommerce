from django.shortcuts import render, redirect

from django.contrib import messages

from django.contrib.auth import authenticate

from django.contrib.auth import login

from django.views.decorators.cache import never_cache

from django.contrib.admin.views.decorators import staff_member_required

from django.contrib.auth import logout

from django.contrib.auth import update_session_auth_hash

from django.utils import timezone

from django.db.models import Sum, Count, Avg

from django.db.models.functions import TruncMonth

from apps.orders.models import Order, OrderItem

from apps.accounts.models import User

import json

from datetime import timedelta


@never_cache
def admin_login(request):

    if request.user.is_authenticated and request.user.is_staff:

        return redirect('admin_dashboard')

    if request.method == 'POST':

        email = request.POST.get('email').strip()

        password = request.POST.get('password')

        user = authenticate(
            request,
            email=email,
            password=password
        )

        if user is not None and user.is_staff:

            login(request,user)

            return redirect('admin_dashboard')

        messages.error(request,'Invalid admin credentials')

    return render(request,'admin/accounts/login.html')



@staff_member_required(login_url='admin_login')
def admin_dashboard(request):
    from apps.products.models import Product
    from django.db.models import Sum, Avg
    from apps.orders.models import Order
    from apps.accounts.models import User
    from django.utils import timezone
    
    today = timezone.now().date()
    first_day_of_month = today.replace(day=1)
    
    today_orders = Order.objects.filter(created_at__date=today)
    today_paid_orders = today_orders.filter(payment_status='paid')
    
    orders_today = today_paid_orders.count()
    
    today_gross = today_paid_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    today_discount = today_paid_orders.aggregate(Sum('coupon_discount'))['coupon_discount__sum'] or 0
    revenue_today = today_gross - today_discount
    
    avg_order_today = today_paid_orders.aggregate(Avg('total_amount'))['total_amount__avg'] or 0
    
    refunds_today = today_orders.filter(order_status='returned').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    all_paid_orders = Order.objects.filter(payment_status='paid')
    
    orders_all = all_paid_orders.count()
    
    gross_all = all_paid_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    discount_all = all_paid_orders.aggregate(Sum('coupon_discount'))['coupon_discount__sum'] or 0
    net_all = gross_all - discount_all
    
    avg_order_all = all_paid_orders.aggregate(Avg('total_amount'))['total_amount__avg'] or 0
    
    refunds_all = Order.objects.filter(order_status='returned').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    customers_all = User.objects.filter(is_staff=False, is_superuser=False).count()
    products_all = Product.objects.count()
    
    month_paid_orders = Order.objects.filter(created_at__date__gte=first_day_of_month, payment_status='paid')
    month_gross = month_paid_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    month_discount = month_paid_orders.aggregate(Sum('coupon_discount'))['coupon_discount__sum'] or 0
    month_revenue = month_gross - month_discount
    
    pending_orders = Order.objects.filter(order_status='pending').count()
    delivered_orders = Order.objects.filter(order_status='delivered').count()
    
    top_seller_today = OrderItem.objects.filter(
        order__created_at__date=today,
        order__payment_status='paid'
    ).values(
        'product_variant__product__product_name'
    ).annotate(
        total_sold=Sum('quantity')
    ).order_by('-total_sold').first()
    
    # Chart Data for Today (Hourly)
    from django.db.models.functions import TruncHour
    hourly_data = (
        today_paid_orders.annotate(hour=TruncHour('created_at'))
        .values('hour')
        .annotate(revenue=Sum('total_amount'), orders=Count('id'))
        .order_by('hour')
    )
    
    # Initialize 24-hour array
    today_chart_revenue = [0] * 24
    today_chart_orders = [0] * 24
    for entry in hourly_data:
        h = entry['hour'].hour
        today_chart_revenue[h] = float(entry['revenue'] or 0)
        today_chart_orders[h] = entry['orders']
    
    import json
    today_chart_revenue_json = json.dumps(today_chart_revenue)
    today_chart_orders_json = json.dumps(today_chart_orders)
    
    context = {
        'revenue_today': round(revenue_today, 2),
        'orders_today': orders_today,
        'refunds_today': round(refunds_today, 2),
        'avg_order_today': round(avg_order_today, 2),
        
        'gross_all': round(gross_all, 2),
        'discount_all': round(discount_all, 2),
        'net_all': round(net_all, 2),
        'refunds_all': round(refunds_all, 2),
        'avg_order_all': round(avg_order_all, 2),
        'orders_all': orders_all,
        'customers_all': customers_all,
        'products_all': products_all,
        
        'month_revenue': round(month_revenue, 2),
        'pending_orders': pending_orders,
        'delivered_orders': delivered_orders,
        'top_seller_today': top_seller_today,
        'today_chart_revenue': today_chart_revenue_json,
        'today_chart_orders': today_chart_orders_json,
    }

    return render(request, 'admin/accounts/dashboard.html', context)


@staff_member_required(login_url='admin_login')
def analytics_dashboard(request):
    

    now = timezone.now()
    today = now.date()
    period = request.GET.get('period', 'monthly')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')

    if period == 'custom' and date_from_str and date_to_str:
        from datetime import datetime
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            date_from = today.replace(day=1)
            date_to = today
    elif period == 'daily':
        date_from = today
        date_to = today
    elif period == 'weekly':
        date_from = today - timedelta(days=6)
        date_to = today
    elif period == 'yearly':
        date_from = today.replace(month=1, day=1)
        date_to = today
    else:
        period = 'monthly'
        date_from = today.replace(day=1)
        date_to = today

    base_qs = Order.objects.filter(
        payment_status='paid',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    )

    total_orders = base_qs.count()
    gross_revenue = base_qs.aggregate(s=Sum('total_amount'))['s'] or 0
    total_coupon_discount = base_qs.aggregate(s=Sum('coupon_discount'))['s'] or 0
    net_revenue = gross_revenue - total_coupon_discount
    avg_order_value = base_qs.aggregate(a=Avg('total_amount'))['a'] or 0

    returned_qs = Order.objects.filter(
        order_status='returned',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    )
    total_refunds = returned_qs.aggregate(s=Sum('total_amount'))['s'] or 0

    new_customers = User.objects.filter(
        is_staff=False,
        date_joined__date__gte=date_from,
        date_joined__date__lte=date_to
    ).count()
    total_customers = User.objects.filter(is_staff=False, is_superuser=False).count()

    all_paid = Order.objects.filter(payment_status='paid')
    monthly_data = (
        all_paid.annotate(month=TruncMonth('created_at'))
        .values('month').annotate(revenue=Sum('total_amount'), orders=Count('id'))
        .order_by('month')
    )
    last_6 = list(monthly_data)[-6:]
    chart_labels = json.dumps([m['month'].strftime('%b %Y') for m in last_6])
    chart_revenue = json.dumps([float(m['revenue']) for m in last_6])
    chart_orders = json.dumps([m['orders'] for m in last_6])

    top_products = (
        OrderItem.objects.filter(
            order__payment_status='paid',
            order__created_at__date__gte=date_from,
            order__created_at__date__lte=date_to
        )
        .values('product_variant__product__product_name')
        .annotate(total_sold=Sum('quantity'), revenue=Sum('total'))
        .order_by('-total_sold')[:5]
    )

    cat_data = (
        OrderItem.objects.filter(
            order__payment_status='paid',
            order__created_at__date__gte=date_from,
            order__created_at__date__lte=date_to
        )
        .values('product_variant__product__category__name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:5]
    )
    max_cat = max((c['total_sold'] for c in cat_data), default=1) or 1
    top_categories = [
        {**c, 'pct': int(c['total_sold'] / max_cat * 100)}
        for c in cat_data
    ]

    status_counts = {
        'pending': base_qs.filter(order_status='pending').count(),
        'processing': base_qs.filter(order_status='processing').count(),
        'shipped': base_qs.filter(order_status='shipped').count(),
        'delivered': base_qs.filter(order_status='delivered').count(),
        'cancelled': Order.objects.filter(order_status='cancelled', payment_status='paid', created_at__date__gte=date_from, created_at__date__lte=date_to).count(),
    }

    payment_breakdown = (
        base_qs.values('payment_method')
        .annotate(count=Count('id'), revenue=Sum('total_amount'))
        .order_by('-count')
    )

    context = {
        'period': period,
        'period_choices': [
            ('daily', 'Daily'),
            ('weekly', 'Weekly'),
            ('monthly', 'Monthly'),
            ('yearly', 'Yearly'),
        ],
        'date_from': date_from,
        'date_to': date_to,
        'total_orders': total_orders,
        'gross_revenue': round(gross_revenue, 2),
        'net_revenue': round(net_revenue, 2),
        'avg_order_value': round(avg_order_value, 2),
        'total_coupon_discount': round(total_coupon_discount, 2),
        'total_refunds': round(total_refunds, 2),
        'new_customers': new_customers,
        'total_customers': total_customers,
        'chart_labels': chart_labels,
        'chart_revenue': chart_revenue,
        'chart_orders': chart_orders,
        'top_products': top_products,
        'top_categories': top_categories,
        'status_counts': status_counts,
        'payment_breakdown': payment_breakdown,
        'orders_list': base_qs.select_related('user', 'address').order_by('-created_at')[:100], # Max 100 on screen to keep fast
    }
    return render(request, 'admin/accounts/analytics.html', context)



@staff_member_required(login_url='admin_login')
def admin_profile(request):

    return render(request, 'admin/accounts/admin_profile.html')


@staff_member_required(login_url='admin_login')
def admin_edit_profile(request):

    if request.method == 'POST':
        username = request.POST.get('username').strip()
        email = request.POST.get('email').strip()
        phone_number = request.POST.get('phone_number').strip()
        
        profile_picture = request.FILES.get('profile_picture')

        request.user.username = username
        request.user.email = email
        
        if hasattr(request.user, 'phone_number'):
            request.user.phone_number = phone_number
            
        if profile_picture and hasattr(request.user, 'profile_image'):
            request.user.profile_image = profile_picture

        request.user.save()

        return redirect('admin_profile')

    return render(request, 'admin/accounts/edit_profile.html')




def admin_logout(request):

    logout(request)

    return redirect('admin_login')


# -------------------------------------------------------------
# Analytics Exports
# -------------------------------------------------------------

def _get_analytics_date_range(request):
    period = request.GET.get('period', 'monthly')
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    today = timezone.now().date()
    
    if period == 'custom' and date_from_str and date_to_str:
        from datetime import datetime
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            date_from = today.replace(day=1)
            date_to = today
    elif period == 'daily':
        date_from = today
        date_to = today
    elif period == 'weekly':
        date_from = today - timedelta(days=6)
        date_to = today
    elif period == 'yearly':
        date_from = today.replace(month=1, day=1)
        date_to = today
    else:
        period = 'monthly'
        date_from = today.replace(day=1)
        date_to = today
        
    return period, date_from, date_to


def _get_analytics_data(date_from, date_to):
    from decimal import Decimal
    paid_qs = Order.objects.filter(
        payment_status='paid',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    )

    total_orders = paid_qs.count()
    gross_revenue = paid_qs.aggregate(s=Sum('total_amount'))['s'] or Decimal('0')
    total_discount = paid_qs.aggregate(s=Sum('coupon_discount'))['s'] or Decimal('0')
    net_revenue = gross_revenue - total_discount

    returned_qs = Order.objects.filter(
        order_status__in=['returned', 'cancelled'],
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    )
    total_refunds = returned_qs.aggregate(s=Sum('total_amount'))['s'] or Decimal('0')

    avg_order_value = paid_qs.aggregate(a=Avg('total_amount'))['a'] or Decimal('0')

    orders_list = paid_qs.select_related('user', 'address').order_by('-created_at')

    top_products = (
        OrderItem.objects.filter(
            order__payment_status='paid',
            order__created_at__date__gte=date_from,
            order__created_at__date__lte=date_to,
        )
        .values('product_variant__product__product_name', 'product_variant__product__category__name')
        .annotate(total_sold=Sum('quantity'), revenue=Sum('total'))
        .order_by('-total_sold')[:10]
    )

    top_categories = (
        OrderItem.objects.filter(
            order__payment_status='paid',
            order__created_at__date__gte=date_from,
            order__created_at__date__lte=date_to,
        )
        .values('product_variant__product__category__name')
        .annotate(total_sold=Sum('quantity'), revenue=Sum('total'))
        .order_by('-total_sold')[:10]
    )

    return {
        'total_orders': total_orders,
        'gross_revenue': round(gross_revenue, 2),
        'total_discount': round(total_discount, 2),
        'net_revenue': round(net_revenue, 2),
        'total_refunds': round(total_refunds, 2),
        'avg_order_value': round(avg_order_value, 2),
        'orders_list': orders_list,
        'top_products': top_products,
        'top_categories': top_categories,
    }


@staff_member_required(login_url='admin_login')
def analytics_pdf(request):
    import io
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    period, date_from, date_to = _get_analytics_date_range(request)
    data = _get_analytics_data(date_from, date_to)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=1.5*cm, leftMargin=1.5*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('title', fontSize=18, fontName='Helvetica-Bold', spaceAfter=4)
    sub_style = ParagraphStyle('sub', fontSize=9, textColor=colors.grey, spaceAfter=12)
    head_style = ParagraphStyle('head', fontSize=11, fontName='Helvetica-Bold', spaceBefore=16, spaceAfter=6)

    elements.append(Paragraph('ShoeDrop — Sales & Analytics Report', title_style))
    elements.append(Paragraph(f'Period: {period.capitalize()}   |   {date_from} to {date_to}', sub_style))

    summary_data = [
        ['Metric', 'Value'],
        ['Total Orders', str(data['total_orders'])],
        ['Gross Revenue', f"Rs. {data['gross_revenue']}"],
        ['Total Discount', f"-Rs. {data['total_discount']}"],
        ['Net Revenue', f"Rs. {data['net_revenue']}"],
        ['Total Refunds', f"Rs. {data['total_refunds']}"],
        ['Avg. Order Value', f"Rs. {data['avg_order_value']}"],
    ]
    summary_table = Table(summary_data, colWidths=[7*cm, 6*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#032b1d')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f4f9f6')]),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#cccccc')),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    elements.append(summary_table)

    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph('Order Details', head_style))
    order_header = ['Order ID', 'Date', 'Customer', 'Payment', 'Discount', 'Status', 'Total']
    order_rows = [order_header]
    for o in data['orders_list']:
        order_rows.append([
            f'#SV-{o.id:04d}',
            o.created_at.strftime('%Y-%m-%d'),
            o.user.username if o.user else '-',
            o.payment_method,
            f"Rs. {o.coupon_discount or 0}",
            o.order_status.replace('_', ' ').title(),
            f"Rs. {o.total_amount}",
        ])
    orders_table = Table(order_rows, colWidths=[2.2*cm, 2.8*cm, 4*cm, 2.8*cm, 2.8*cm, 3.2*cm, 3*cm])
    orders_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#032b1d')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f4f9f6')]),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#cccccc')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(orders_table)

    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph('Top 10 Best Selling Products', head_style))
    prod_header = ['Rank', 'Product', 'Category', 'Units Sold', 'Revenue']
    prod_rows = [prod_header]
    for i, p in enumerate(data['top_products'], 1):
        prod_rows.append([
            str(i),
            p['product_variant__product__product_name'] or '-',
            p['product_variant__product__category__name'] or '-',
            str(p['total_sold']),
            f"Rs. {p['revenue']}",
        ])
    prod_table = Table(prod_rows, colWidths=[1.5*cm, 8*cm, 5*cm, 3*cm, 3.5*cm])
    prod_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#032b1d')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f4f9f6')]),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#cccccc')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(prod_table)

    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="analytics_report_{date_from}_{date_to}.pdf"'
    return response


@staff_member_required(login_url='admin_login')
def analytics_excel(request):
    import openpyxl
    import io
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    period, date_from, date_to = _get_analytics_date_range(request)
    data = _get_analytics_data(date_from, date_to)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Analytics Report'

    green_fill = PatternFill('solid', fgColor='032B1D')
    light_fill = PatternFill('solid', fgColor='F4F9F6')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    title_font = Font(name='Calibri', bold=True, color='032B1D', size=13)
    bold_font = Font(name='Calibri', bold=True, size=10)
    normal_font = Font(name='Calibri', size=10)
    thin = Side(style='thin', color='CCCCCC')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:D1')
    ws['A1'] = f'ShoeDrop Analytics Report — {period.capitalize()} | {date_from} to {date_to}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = green_fill
    ws.row_dimensions[1].height = 28

    # --- SUMMARY ---
    summary_rows = [
        ('Metric', 'Value'),
        ('Total Orders', data['total_orders']),
        ('Gross Revenue', f"Rs. {data['gross_revenue']}"),
        ('Total Discount', f"Rs. {data['total_discount']}"),
        ('Net Revenue', f"Rs. {data['net_revenue']}"),
        ('Total Refunds', f"Rs. {data['total_refunds']}"),
        ('Avg. Order Value', f"Rs. {data['avg_order_value']}"),
    ]
    current_row = 3
    for row_idx, (label, value) in enumerate(summary_rows, start=current_row):
        ws.cell(row_idx, 1, label).font = bold_font if row_idx == current_row else normal_font
        ws.cell(row_idx, 2, value).font = bold_font if row_idx == current_row else normal_font
        if row_idx == current_row:
            ws.cell(row_idx, 1).fill = green_fill
            ws.cell(row_idx, 2).fill = green_fill
            ws.cell(row_idx, 1).font = header_font
            ws.cell(row_idx, 2).font = header_font
        elif row_idx % 2 == 0:
            ws.cell(row_idx, 1).fill = light_fill
            ws.cell(row_idx, 2).fill = light_fill
        for col in (1, 2):
            ws.cell(row_idx, col).border = thin_border
            ws.cell(row_idx, col).alignment = Alignment(vertical='center', horizontal='left')
    
    current_row += len(summary_rows) + 2

    # --- ORDERS ---
    ws.cell(current_row, 1, 'Order Details').font = title_font
    current_row += 1
    
    order_headers = ['Order ID', 'Date', 'Customer', 'Email', 'Payment Method', 'Discount', 'Status', 'Total']
    for col_idx, h in enumerate(order_headers, 1):
        cell = ws.cell(current_row, col_idx, h)
        cell.font = header_font
        cell.fill = green_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    for row_idx, o in enumerate(data['orders_list'], start=current_row):
        row_data = [
            f'#SV-{o.id:04d}',
            o.created_at.strftime('%Y-%m-%d %H:%M'),
            o.user.username if o.user else '-',
            o.user.email if o.user else '-',
            o.payment_method,
            f"Rs. {o.coupon_discount or 0}",
            o.order_status.replace('_', ' ').title(),
            f"Rs. {o.total_amount}",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.font = normal_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = light_fill
    
    current_row += len(data['orders_list']) + 2

    # --- TOP PRODUCTS ---
    ws.cell(current_row, 1, 'Top 10 Best Selling Products').font = title_font
    current_row += 1
    
    prod_headers = ['Rank', 'Product Name', 'Category', 'Units Sold', 'Revenue']
    for col_idx, h in enumerate(prod_headers, 1):
        cell = ws.cell(current_row, col_idx, h)
        cell.font = header_font
        cell.fill = green_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    for row_idx, p in enumerate(data['top_products'], start=current_row):
        row_data = [
            row_idx - current_row + 1,
            p['product_variant__product__product_name'] or '-',
            p['product_variant__product__category__name'] or '-',
            p['total_sold'],
            f"Rs. {p['revenue']}",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.font = normal_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = light_fill
    
    # Set Column Widths for better reading
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="analytics_report_{date_from}_{date_to}.xlsx"'
    return response